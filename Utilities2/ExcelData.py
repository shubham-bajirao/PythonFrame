import openpyxl

#print(sheet.cell(row=2, column=3).value)
#print(sheet.max_row)
class Excelvalue:

    @staticmethod
    def ExcelValues():
        book = openpyxl.load_workbook("C:\\Users\\LENOVO\\PycharmProjects\\Book1.xlsx")
        list1 = []
        sheet = book.active
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == "TestCase2":
                for j in range(2, sheet.max_column + 1):
                    list1.append(sheet.cell(row=i, column=j).value)
        return(list1)